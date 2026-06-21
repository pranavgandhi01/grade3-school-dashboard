#!/usr/bin/env python3
"""
Incremental Knowledge Graph Sync Engine for Grade 3 School Dashboard.

This script maintains a persistent knowledge graph that tracks all school data.
It uses file hashing to detect changes and only reprocesses modified/new files.

Usage:
  python3 scripts/sync.py              # Incremental sync (only new/changed files)
  python3 scripts/sync.py --full       # Full re-sync (reprocess everything)
  python3 scripts/sync.py --status     # Show what's tracked and what's new

Data Sources:
  - ~/Desktop/Grade 3/docs/        (daily updates, circulars, schedules)
  - ~/Desktop/Grade 3/pics/        (WhatsApp images — activity flyers)
  - ~/Downloads/docs/Annexure*.pdf (school annexures — holidays, menus, etc.)
"""

import hashlib
import json
import os
import re
import sys
import glob
from datetime import datetime, date
from pathlib import Path

# --- Configuration ---
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data"
MANIFEST_PATH = DATA_DIR / "manifest.json"
KNOWLEDGE_GRAPH_PATH = DATA_DIR / "knowledge_graph.json"
SCHOOL_DATA_JS_PATH = PROJECT_ROOT / "src" / "data" / "schoolData.js"

SOURCE_DIRS = [
    Path.home() / "Desktop" / "Grade 3" / "docs",
    Path.home() / "Desktop" / "Grade 3" / "pics",
]
SOURCE_GLOBS = [
    str(Path.home() / "Downloads" / "docs" / "Annexure*.pdf"),
]

SUPPORTED_EXTENSIONS = {".pdf", ".xlsx", ".xls", ".jpeg", ".jpg", ".png", ".mp4"}


# --- File Hashing ---
def file_hash(filepath: str) -> str:
    """SHA256 hash of file contents for change detection."""
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


# --- Manifest Management ---
def load_manifest() -> dict:
    if MANIFEST_PATH.exists():
        with open(MANIFEST_PATH) as f:
            return json.load(f)
    return {"version": 1, "files": {}, "last_sync": None}


def save_manifest(manifest: dict):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    manifest["last_sync"] = datetime.now().isoformat()
    with open(MANIFEST_PATH, "w") as f:
        json.dump(manifest, f, indent=2)


# --- Knowledge Graph Management ---
def load_knowledge_graph() -> dict:
    if KNOWLEDGE_GRAPH_PATH.exists():
        with open(KNOWLEDGE_GRAPH_PATH) as f:
            return json.load(f)
    return {
        "version": 1,
        "school": "VIBGYOR Rise Malad",
        "grade": "3B",
        "academic_year": "2026-27",
        "last_updated": None,
        "holidays": [],
        "events": [],
        "daily_updates": [],
        "review_schedule": [],
        "dictation_words": [],
        "cafeteria_menu": [],
        "spa_activities": [],
        "external_exams": [],
        "dialogues": [],
        "study_materials": [],
        "circulars": [],
        "processed_files": [],
    }


def save_knowledge_graph(kg: dict):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    kg["last_updated"] = datetime.now().isoformat()
    with open(KNOWLEDGE_GRAPH_PATH, "w") as f:
        json.dump(kg, f, indent=2, default=str, ensure_ascii=False)


# --- PDF Text Extraction ---
def extract_pdf_text(filepath: str) -> dict:
    """Extract text and tables from a PDF file."""
    import pdfplumber
    result = {"text": "", "tables": [], "pages": 0}
    try:
        with pdfplumber.open(filepath) as pdf:
            pages_text = []
            for page in pdf.pages:
                text = page.extract_text() or ""
                pages_text.append(text)
                page_tables = page.extract_tables()
                if page_tables:
                    for t in page_tables:
                        result["tables"].append(t)
            result["text"] = "\n".join(pages_text)
            result["pages"] = len(pdf.pages)
    except Exception as e:
        result["error"] = str(e)
    return result


def extract_excel(filepath: str) -> list:
    """Extract rows from Excel file."""
    import openpyxl
    rows_out = []
    try:
        wb = openpyxl.load_workbook(filepath)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                row_data = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in row_data):
                    rows_out.append(row_data)
    except Exception as e:
        print(f"  ⚠ Excel error: {e}")
    return rows_out


# --- Content Classifiers ---
def classify_file(filename: str) -> str:
    """Classify a file by its content type based on filename patterns."""
    fn = filename.upper()
    if "DAILY UPDATE" in fn:
        return "daily_update"
    elif "HOLIDAY" in fn:
        return "holidays"
    elif "TERM" in fn and "CALENDAR" in fn:
        return "term_calendar"
    elif "REVIEW SCHEDULE" in fn:
        return "review_schedule"
    elif "CAFETERIA" in fn or "MENU" in fn:
        return "cafeteria_menu"
    elif "FIELD TRIP" in fn:
        return "field_trip"
    elif "PICNIC" in fn or "ONE DAY" in fn:
        return "picnic"
    elif "IHC" in fn or "INTERHOUSE" in fn:
        return "ihc"
    elif "FUN FRIDAY" in fn:
        return "fun_friday"
    elif "PTM" in fn:
        return "ptm"
    elif "EXTERNAL EXAM" in fn:
        return "external_exams"
    elif "DIALOGUE" in fn:
        return "dialogues"
    elif "FOOTBALL" in fn or "CRICKET" in fn:
        return "sports_circular"
    elif "GREENFLUENCER" in fn:
        return "activity"
    elif "PARENT APP" in fn:
        return "parent_app_manual"
    elif "PSA" in fn or "ROLEPLAY" in fn:
        return "study_material"
    elif fn.endswith((".JPEG", ".JPG", ".PNG")):
        return "image_flyer"
    elif fn.endswith(".MP4"):
        return "video"
    else:
        return "other"


# --- Data Parsers ---
def parse_daily_update(text: str, filename: str) -> dict:
    """Parse a daily update PDF into structured periods."""
    # Extract date
    date_match = re.search(r"DATE\s*[-–:]\s*(\d{1,2}\s+\w+\s+\d{4})", text)
    update_date = date_match.group(1).strip() if date_match else filename

    # Extract grade
    grade_match = re.search(r"Grade\s*[-–:]\s*(\w+)", text)
    grade = grade_match.group(1).strip() if grade_match else "3B"

    # Extract periods
    periods = []
    period_blocks = re.split(r"Period\s*[-–:]\s*(\d+)", text)

    for i in range(1, len(period_blocks) - 1, 2):
        period_num = int(period_blocks[i])
        block = period_blocks[i + 1]

        def extract_field(name, block_text):
            match = re.search(
                rf"{name}\s+(.*?)(?=(?:Subject|Topic|Sub Topic|CW|Reinforcement|Submission date|Period|New Words|Additional)\s|$)",
                block_text,
                re.DOTALL,
            )
            if match:
                val = match.group(1).strip()
                # Clean up multi-line
                val = re.sub(r"\s+", " ", val)
                return val if val and val.upper() != "NIL" else None
            return None

        subject = extract_field("Subject", block)
        topic = extract_field("Topic", block)
        sub_topic = extract_field("Sub Topic", block)
        cw = extract_field("CW", block)
        reinforcement = extract_field("Reinforcement", block)
        submission = extract_field("Submission date", block)

        if subject:
            periods.append({
                "period": period_num,
                "subject": subject,
                "topic": topic,
                "sub_topic": sub_topic,
                "classwork": cw,
                "homework": reinforcement,
                "submission_date": submission,
            })

    # Extract new words
    new_words_match = re.search(r"New Words\s*(.*?)(?=Additional Information|$)", text, re.DOTALL)
    new_words = []
    if new_words_match:
        words_text = new_words_match.group(1).strip()
        new_words = [w.strip() for w in re.split(r"[,\n]", words_text) if w.strip()]

    # Extract additional information
    add_info_match = re.search(r"Additional Information\s*(.*?)$", text, re.DOTALL)
    additional_info = add_info_match.group(1).strip() if add_info_match else None
    if additional_info and not additional_info:
        additional_info = None

    return {
        "date": update_date,
        "grade": grade,
        "periods": periods,
        "new_words": new_words,
        "additional_info": additional_info,
        "source_file": filename,
    }


def parse_holidays(text: str) -> list:
    """Parse holiday list into structured entries."""
    holidays = []
    lines = text.split("\n")
    for line in lines:
        # Match patterns like: 1 GOOD FRIDAY 3-Apr-26 Friday 3-Apr-26 Friday 1
        match = re.search(
            r"(\d+)\s+(.+?)\s+(\d{1,2}-\w{3}-\d{2})\s+\w+\s+(\d{1,2}-\w{3}-\d{2})\s+\w+\s+(\d+)",
            line,
        )
        if match:
            holidays.append({
                "sr_no": int(match.group(1)),
                "name": match.group(2).strip(),
                "from_date": match.group(3),
                "to_date": match.group(4),
                "days": int(match.group(5)),
            })
    return holidays


def parse_review_schedule(tables: list) -> list:
    """Parse review schedule tables."""
    schedule = []
    for table in tables:
        for row in table:
            if not row or not row[0] or row[0] == "SUBJECT" or "Grade 3" in str(row[0]):
                continue
            cells = [str(c) if c else "" for c in row]
            if len(cells) >= 6 and cells[0].strip():
                entry = {
                    "subject": cells[0].strip(),
                    "periodic_test_date": cells[1].strip() if cells[1].strip() != "NIL" else None,
                    "periodic_test_topic": cells[2].strip() if cells[2].strip() != "NIL" else None,
                    "sea_date": cells[3].strip() if len(cells) > 3 and cells[3].strip() != "NIL" else None,
                    "sea_topic": cells[4].strip() if len(cells) > 4 and cells[4].strip() != "NIL" else None,
                    "sea_details": cells[5].strip() if len(cells) > 5 and cells[5].strip() != "NIL" else None,
                }
                if entry["periodic_test_date"] or entry["sea_date"]:
                    schedule.append(entry)
    return schedule


def parse_cafeteria_menu(text: str) -> list:
    """Parse cafeteria menu into weekly structure. Simplified extraction."""
    weeks = []
    # The menu is complex — store as raw text with week identifiers
    week_blocks = re.split(r"Morning Snacks\s+(\d+)", text)
    seen_weeks = set()
    for i in range(1, len(week_blocks) - 1, 2):
        week_num = week_blocks[i]
        if week_num not in seen_weeks:
            seen_weeks.add(week_num)
            weeks.append({
                "week": int(week_num),
                "content": week_blocks[i + 1][:300].strip(),
            })
    return weeks


def parse_spa_activities(rows: list) -> list:
    """Parse PSA Plan Excel data."""
    activities = []
    seen = set()
    header = rows[0] if rows else []
    for row in rows[1:]:
        if len(row) >= 3 and row[1].strip() and row[1].strip() not in seen:
            activity_name = row[1].strip()
            seen.add(activity_name)
            # Find timing columns
            schedule = {}
            days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
            for d_idx, day in enumerate(days):
                col_start = 6 + d_idx * 2
                col_end = 7 + d_idx * 2
                if col_start < len(row) and row[col_start].strip():
                    schedule[day] = {
                        "start": row[col_start].strip(),
                        "end": row[col_end].strip() if col_end < len(row) else "",
                    }

            fee_col = -2 if len(row) >= 2 else -1
            fee = ""
            for c in row[-3:]:
                if c.strip() and c.strip().isdigit():
                    fee = c.strip()
                    break

            activities.append({
                "name": activity_name,
                "grades": f"{row[2]} - {row[3]}" if len(row) > 3 else "",
                "schedule": schedule,
                "fee_half_yearly": fee,
            })
    return activities


# --- Knowledge Graph Merger ---
def merge_into_kg(kg: dict, file_type: str, data: dict, filename: str):
    """Merge extracted data into the knowledge graph without duplicates."""

    if file_type == "daily_update":
        parsed = parse_daily_update(data["text"], filename)
        # Remove existing entry for same date (in case of update)
        kg["daily_updates"] = [
            u for u in kg["daily_updates"] if u.get("date") != parsed["date"]
        ]
        kg["daily_updates"].append(parsed)
        # Merge new words into dictation repository
        for word in parsed.get("new_words", []):
            existing = {w["word"].lower() for w in kg["dictation_words"]}
            if word.lower() not in existing:
                kg["dictation_words"].append({
                    "word": word,
                    "date": parsed["date"],
                    "source": filename,
                })

    elif file_type == "holidays":
        holidays = parse_holidays(data["text"])
        if holidays:
            kg["holidays"] = holidays  # Replace entirely — single source of truth

    elif file_type == "review_schedule":
        schedule = parse_review_schedule(data.get("tables", []))
        if schedule:
            kg["review_schedule"] = schedule

    elif file_type == "cafeteria_menu":
        kg["cafeteria_menu"] = {
            "raw_text": data["text"],
            "tables": data.get("tables", []),
        }

    elif file_type == "term_calendar":
        kg["events"] = [e for e in kg["events"] if e.get("source_type") != "term_calendar"]
        # Parse key events from term calendar text
        text = data["text"]
        term_events = []
        event_patterns = [
            (r"PTM\s*1", "PTM 1", "25-Apr-2026"),
            (r"PTM\s*2.*?Morning", "PTM 2 (Morning)", "01-Aug-2026"),
            (r"PTM\s*2.*?Afternoon", "PTM 2 (Afternoon)", "08-Aug-2026"),
            (r"Field Trip\s*[-–]\s*1", "Field Trip 1 - Gymnasium", "23-Apr-2026"),
            (r"Field Trip\s*[-–]\s*2", "Field Trip 2", "11-Sep-2026"),
            (r"IHC\s*1", "Inter House Competition 1 - Paper Puppet", "25-Jun-2026"),
            (r"IHC\s*2", "Inter House Competition 2", "11-Aug-2026"),
            (r"IHC\s*3", "Inter House Competition 3", "29-Sep-2026"),
            (r"One Day Trip", "One Day Picnic - Smaaash", "23-Jul-2026"),
            (r"Food Festival", "Food Festival", "07-Sep-2026"),
            (r"Term End Party", "Culminating Term 1 / Term End Party", "30-Sep-2026"),
            (r"Fun Friday", "Fun Friday", "19-Jun-2026"),
        ]
        for pattern, name, date_str in event_patterns:
            if re.search(pattern, text, re.IGNORECASE):
                term_events.append({
                    "name": name,
                    "date": date_str,
                    "source_type": "term_calendar",
                    "category": "school_event",
                })
        kg["events"].extend(term_events)

    elif file_type == "field_trip":
        event = {
            "name": "Field Trip 1 - Gymnasium Visit",
            "date": "23-Apr-2026",
            "details": data["text"][:500],
            "source_type": "circular",
            "category": "field_trip",
            "things_to_carry": "Cap, water bottle, handkerchief, snacks, sanitiser, mask",
            "uniform": "SPA uniform with ID Card",
        }
        kg["events"] = [e for e in kg["events"] if e.get("name") != event["name"] or e.get("source_type") != "circular"]
        kg["events"].append(event)

    elif file_type == "picnic":
        event = {
            "name": "One Day Picnic - Smaaash, Lower Parel",
            "date": "23-Jul-2026",
            "details": data["text"][:800],
            "source_type": "circular",
            "category": "picnic",
            "cost": "₹1920 per student",
            "reporting_time": "08:00 AM",
            "travel": "AC Luxury Coaches",
        }
        kg["events"] = [e for e in kg["events"] if "Smaaash" not in e.get("name", "") and "Picnic" not in e.get("name", "")]
        kg["events"].append(event)

    elif file_type == "ihc":
        event = {
            "name": "Inter House Competition 1 - Paper Stick Puppet",
            "date": "25-Jun-2026",
            "details": data["text"][:500],
            "source_type": "circular",
            "category": "competition",
            "materials": "Ice cream sticks, cardboard/chart paper",
        }
        kg["events"] = [e for e in kg["events"] if "Inter House Competition 1" not in e.get("name", "")]
        kg["events"].append(event)

    elif file_type == "fun_friday":
        event = {
            "name": "Fun Friday",
            "date": "19-Jun-2026",
            "details": "Students dress in smart casuals. Last Friday of every month.",
            "source_type": "circular",
            "category": "school_event",
            "recurring": "Last Friday of every month",
        }
        kg["events"] = [e for e in kg["events"] if not (e.get("name") == "Fun Friday" and e.get("source_type") == "circular")]
        kg["events"].append(event)

    elif file_type == "ptm":
        event = {
            "name": "Parent Teacher Meeting 1",
            "date": "25-Apr-2026",
            "time": "7:30 AM to 12:30 PM",
            "details": data["text"][:400],
            "source_type": "circular",
            "category": "ptm",
        }
        kg["events"] = [e for e in kg["events"] if not (e.get("name") == event["name"] and e.get("source_type") == "circular")]
        kg["events"].append(event)

    elif file_type == "external_exams":
        kg["external_exams"] = {
            "info": data["text"][:2000],
            "registration_deadline": "30 July 2026",
            "portal": "OSEHUB.COM → OSE HUB under parent services",
            "contact": "600 3000 700",
            "exams": [
                {"name": "LogiQids IIT Bombay", "type": "Logical Reasoning & Computational Thinking", "grades": "All"},
                {"name": "GREEN Olympiad (TERI)", "type": "Environment Quiz", "grades": "All"},
                {"name": "Dr. Homi Bhabha Balvaidnyanik", "type": "Science Competition", "grades": "6th and 9th only"},
            ],
        }

    elif file_type == "dialogues":
        kg["dialogues"] = {
            "title": "The Weightlifting Princess - Roleplay",
            "text": data["text"],
            "source": filename,
        }

    elif file_type == "sports_circular":
        kg["circulars"].append({
            "title": filename.replace(".pdf", ""),
            "text": data["text"][:500],
            "source": filename,
            "type": "sports",
        })

    elif file_type == "study_material":
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            spa = parse_spa_activities(data.get("tables", [[]])[0] if data.get("tables") else [])
            if spa:
                kg["spa_activities"] = spa
        else:
            kg["study_materials"].append({
                "title": filename.replace(".pdf", ""),
                "text": data["text"][:1000] if data.get("text") else "",
                "source": filename,
            })

    elif file_type == "image_flyer":
        # Store metadata about the image
        kg["circulars"].append({
            "title": filename,
            "type": "image_flyer",
            "source": filename,
            "description": "WhatsApp activity flyer image",
        })

    # Track processed file
    if filename not in kg["processed_files"]:
        kg["processed_files"].append(filename)


# --- Source File Discovery ---
def discover_source_files() -> list:
    """Find all source files from configured directories."""
    files = []
    for src_dir in SOURCE_DIRS:
        if src_dir.exists():
            for f in sorted(src_dir.iterdir()):
                if f.suffix.lower() in SUPPORTED_EXTENSIONS:
                    files.append(str(f))

    for pattern in SOURCE_GLOBS:
        for f in sorted(glob.glob(pattern)):
            files.append(f)

    # Also check for PSA Plan Excel
    psa_path = Path.home() / "Desktop" / "Grade 3" / "docs" / "PSA Plan.xlsx"
    if psa_path.exists() and str(psa_path) not in files:
        files.append(str(psa_path))

    return files


# --- Main Sync Logic ---
def sync(full: bool = False):
    """Run incremental (or full) sync of school data into knowledge graph."""
    manifest = load_manifest()
    kg = load_knowledge_graph()
    files = discover_source_files()

    if not files:
        print("❌ No source files found. Check that directories exist:")
        for d in SOURCE_DIRS:
            print(f"   {d} {'✓' if d.exists() else '✗'}")
        return

    new_count = 0
    changed_count = 0
    skipped_count = 0

    print(f"🔍 Found {len(files)} source files")
    print(f"📊 Knowledge graph has {len(kg.get('processed_files', []))} processed files")
    print()

    for filepath in files:
        filename = os.path.basename(filepath)
        current_hash = file_hash(filepath)
        prev_entry = manifest["files"].get(filepath)

        if not full and prev_entry and prev_entry["hash"] == current_hash:
            skipped_count += 1
            continue

        is_new = prev_entry is None
        if is_new:
            new_count += 1
            print(f"  🆕 {filename}")
        else:
            changed_count += 1
            print(f"  🔄 {filename} (changed)")

        # Extract content
        file_type = classify_file(filename)
        ext = os.path.splitext(filename)[1].lower()

        data = {}
        if ext == ".pdf":
            data = extract_pdf_text(filepath)
        elif ext in (".xlsx", ".xls"):
            rows = extract_excel(filepath)
            data = {"tables": [rows]}
            file_type = "study_material"  # PSA Plan
        elif ext in (".jpeg", ".jpg", ".png"):
            data = {"type": "image", "path": filepath}
        elif ext == ".mp4":
            data = {"type": "video", "path": filepath}

        # Merge into knowledge graph
        merge_into_kg(kg, file_type, data, filename)

        # Update manifest
        manifest["files"][filepath] = {
            "hash": current_hash,
            "filename": filename,
            "file_type": file_type,
            "size": os.path.getsize(filepath),
            "last_processed": datetime.now().isoformat(),
            "mod_time": datetime.fromtimestamp(os.path.getmtime(filepath)).isoformat(),
        }

    def sort_date(date_str):
        if not date_str:
            return datetime.min
        try:
            return datetime.strptime(date_str.strip().upper(), "%d %B %Y")
        except:
            return datetime.min

    # Sort daily updates by date
    kg["daily_updates"].sort(key=lambda x: sort_date(x.get("date", "")))

    # Sort dictation words by date
    kg["dictation_words"].sort(key=lambda x: sort_date(x.get("date", "")))

    # Deduplicate circulars
    seen_circulars = set()
    unique_circulars = []
    for c in kg["circulars"]:
        key = c.get("source", c.get("title", ""))
        if key not in seen_circulars:
            seen_circulars.add(key)
            unique_circulars.append(c)
    kg["circulars"] = unique_circulars

    # Deduplicate events (prefer circulars over term_calendar)
    # If a circular exists for an event, we remove the term_calendar equivalent.
    circular_events = [e for e in kg.get("events", []) if e.get("source_type") == "circular"]
    final_events = list(circular_events)
    
    for term_e in [e for e in kg.get("events", []) if e.get("source_type") == "term_calendar"]:
        # Check if this term event is superseded by any circular event
        is_superseded = False
        for circ_e in circular_events:
            # Match by prefix/keywords
            if "Inter House Competition 1" in term_e["name"] and "Inter House Competition 1" in circ_e["name"]:
                is_superseded = True
                break
            if "Picnic" in term_e["name"] and ("Picnic" in circ_e["name"] or "Smaaash" in circ_e["name"]):
                is_superseded = True
                break
            if "Field Trip 1" in term_e["name"] and "Field Trip 1" in circ_e["name"]:
                is_superseded = True
                break
            if "Fun Friday" in term_e["name"] and "Fun Friday" in circ_e["name"]:
                is_superseded = True
                break
            if "PTM 1" in term_e["name"] and "PTM 1" in circ_e["name"]:
                is_superseded = True
                break
        
        if not is_superseded:
            final_events.append(term_e)
            
    kg["events"] = final_events

    # Save
    save_manifest(manifest)
    save_knowledge_graph(kg)

    # Generate JS data module
    generate_js_data(kg)

    print()
    print(f"✅ Sync complete:")
    print(f"   🆕 New:     {new_count}")
    print(f"   🔄 Changed: {changed_count}")
    print(f"   ⏭️  Skipped: {skipped_count} (unchanged)")
    print(f"   📊 Knowledge graph: {KNOWLEDGE_GRAPH_PATH}")
    print(f"   📱 React data:      {SCHOOL_DATA_JS_PATH}")
    print(f"   📅 Holidays: {len(kg.get('holidays', []))}")
    print(f"   📝 Daily Updates: {len(kg.get('daily_updates', []))}")
    print(f"   📖 Dictation Words: {len(kg.get('dictation_words', []))}")
    print(f"   🎯 Events: {len(kg.get('events', []))}")
    print(f"   📋 Review Schedule: {len(kg.get('review_schedule', []))}")


def show_status():
    """Show current sync status."""
    manifest = load_manifest()
    kg = load_knowledge_graph()
    files = discover_source_files()

    tracked = set(manifest["files"].keys())
    current = set(files)

    new_files = current - tracked
    removed = tracked - current
    changed = []

    for f in tracked & current:
        if file_hash(f) != manifest["files"][f]["hash"]:
            changed.append(f)

    print(f"📊 Knowledge Graph Status")
    print(f"   Last sync: {manifest.get('last_sync', 'Never')}")
    print(f"   Tracked files: {len(tracked)}")
    print(f"   Source files: {len(current)}")
    print()

    if new_files:
        print(f"🆕 New files ({len(new_files)}):")
        for f in sorted(new_files):
            print(f"   + {os.path.basename(f)}")

    if changed:
        print(f"🔄 Changed files ({len(changed)}):")
        for f in changed:
            print(f"   ~ {os.path.basename(f)}")

    if removed:
        print(f"❌ Removed files ({len(removed)}):")
        for f in sorted(removed):
            print(f"   - {os.path.basename(f)}")

    if not new_files and not changed and not removed:
        print("✅ Everything is up to date!")

    print()
    print(f"📅 Holidays: {len(kg.get('holidays', []))}")
    print(f"📝 Daily Updates: {len(kg.get('daily_updates', []))}")
    print(f"📖 Dictation Words: {len(kg.get('dictation_words', []))}")
    print(f"🎯 Events: {len(kg.get('events', []))}")
    print(f"📋 Review Schedule: {len(kg.get('review_schedule', []))}")
    print(f"🏅 SPA Activities: {len(kg.get('spa_activities', []))}")


# --- JS Data Generator ---
def generate_js_data(kg: dict):
    """Generate the React-consumable schoolData.js module from knowledge graph."""
    SCHOOL_DATA_JS_PATH.parent.mkdir(parents=True, exist_ok=True)

    empty_dict = {}
    empty_list = []

    parts = []
    parts.append(f"// Auto-generated by scripts/sync.py — DO NOT EDIT MANUALLY")
    parts.append(f"// Last synced: {datetime.now().isoformat()}")
    parts.append(f"// Source: Knowledge Graph v{kg.get('version', 1)}")
    parts.append("")

    school = kg.get('school', 'VIBGYOR Rise Malad')
    grade = kg.get('grade', '3B')
    ay = kg.get('academic_year', '2026-27')
    parts.append(f'export const SCHOOL_INFO = {{')
    parts.append(f'  school: "{school}",')
    parts.append(f'  grade: "{grade}",')
    parts.append(f'  academicYear: "{ay}",')
    parts.append(f'  address: "Vibgyor Rise Malad, Ram Nagar, Malad West, Mumbai, Maharashtra 400064",')
    parts.append(f'  phone: "02249119600",')
    parts.append(f'  parentApp: "Hubble Orion Parent App",')
    parts.append(f'  portal: "www.osehub.com",')
    parts.append(f'}};')
    parts.append("")

    def dump(data):
        return json.dumps(data, indent=2, ensure_ascii=False, default=str)

    parts.append(f"export const HOLIDAYS = {dump(kg.get('holidays', empty_list))};")
    parts.append("")
    parts.append(f"export const EVENTS = {dump(kg.get('events', empty_list))};")
    parts.append("")
    parts.append(f"export const DAILY_UPDATES = {dump(kg.get('daily_updates', empty_list))};")
    parts.append("")
    parts.append(f"export const REVIEW_SCHEDULE = {dump(kg.get('review_schedule', empty_list))};")
    parts.append("")
    parts.append(f"export const DICTATION_WORDS = {dump(kg.get('dictation_words', empty_list))};")
    parts.append("")
    parts.append(f"export const CAFETERIA_MENU = {dump(kg.get('cafeteria_menu', empty_dict))};")
    parts.append("")
    parts.append(f"export const SPA_ACTIVITIES = {dump(kg.get('spa_activities', empty_list))};")
    parts.append("")
    parts.append(f"export const EXTERNAL_EXAMS = {dump(kg.get('external_exams', empty_dict))};")
    parts.append("")
    parts.append(f"export const DIALOGUES = {dump(kg.get('dialogues', empty_dict))};")
    parts.append("")
    parts.append(f"export const CIRCULARS = {dump(kg.get('circulars', empty_list))};")
    parts.append("")
    parts.append(f"export const STUDY_MATERIALS = {dump(kg.get('study_materials', empty_list))};")
    parts.append("")

    with open(SCHOOL_DATA_JS_PATH, "w") as f:
        f.write("\n".join(parts))


# --- CLI ---
if __name__ == "__main__":
    if "--status" in sys.argv:
        show_status()
    elif "--full" in sys.argv:
        print("🔄 Full re-sync (reprocessing all files)...\\n")
        sync(full=True)
    else:
        print("🔄 Incremental sync (only new/changed files)...\\n")
        sync(full=False)
